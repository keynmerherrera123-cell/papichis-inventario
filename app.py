from flask import Flask, render_template_string, request, redirect, url_for, session, send_file
import sqlite3
import io
from datetime import datetime

# Librerías de OpenPyXL para ponerle estilo al Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

app = Flask(__name__)
app.secret_key = 'clave_secreta_super_segura_papichis'


# --- ZONA DE TRABAJADORES (Vista Móvil Básica) ---

@app.route('/')
def mostrar_inventario():
    conn = sqlite3.connect('inventario_papichis.db')
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre_producto, cantidad_actual, categoria FROM productos")
    datos = cursor.fetchall()
    conn.close()

    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Control Papichis</title>
        <style>
            body { font-family: Arial; padding: 15px; background: #f4f4f4; margin: 0; }
            h2 { color: #333; text-align: center; }
            .tarjeta { background: white; padding: 15px; margin-bottom: 12px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); display: flex; justify-content: space-between; align-items: center;}
            .info { flex-grow: 1; }
            .formulario-edicion { display: flex; gap: 8px; }
            input[type="number"] { width: 60px; padding: 8px; font-size: 16px; text-align: center; border: 1px solid #ccc; border-radius: 4px; }
            button { background: #007bff; color: white; border: none; padding: 8px 15px; border-radius: 4px; font-weight: bold; cursor: pointer; }
        </style>
    </head>
    <body>
        <h2>📱 Inventario Papichis</h2>
        {% for producto in datos %}
        <div class="tarjeta">
            <div class="info">
                <strong>{{ producto[1] }}</strong><br>
                <small style="color: gray;">{{ producto[3] }}</small>
            </div>
            <form action="/actualizar" method="POST" class="formulario-edicion">
                <input type="hidden" name="id_producto" value="{{ producto[0] }}">
                <input type="number" name="nueva_cantidad" value="{{ producto[2] }}" min="0">
                <button type="submit">Guardar</button>
            </form>
        </div>
        {% endfor %}
        <br>
        <div style="text-align: center;">
            <a href="/admin" style="color: gray; text-decoration: none; font-size: 12px;">Acceso Administrador</a>
        </div>
    </body>
    </html>
    """
    return render_template_string(html, datos=datos)

@app.route('/actualizar', methods=['POST'])
def actualizar_inventario():
    producto_id = request.form['id_producto']
    nueva_cantidad = request.form['nueva_cantidad']
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    conn = sqlite3.connect('inventario_papichis.db')
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE productos SET cantidad_actual = ?, ultima_actualizacion = ? WHERE id = ?
    ''', (nueva_cantidad, fecha_actual, producto_id))
    conn.commit()
    conn.close()
    return redirect(url_for('mostrar_inventario'))


# --- ZONA DE ADMINISTRACIÓN (Dashboard con Analíticas Avanzadas) ---

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        clave = request.form.get('clave')
        if clave == 'papichis2026':
            session['admin_autenticado'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            return "<h2 style='color:red; text-align:center;'>Contraseña incorrecta. <a href='/admin'>Volver</a></h2>"
    
    html = """
    <!DOCTYPE html>
    <html>
    <head><title>Login Papichis</title></head>
    <body style="font-family: Arial; text-align: center; padding-top: 100px; background: #333; color: white;">
        <h2>Acceso Exclusivo Administrador</h2>
        <form action="/admin" method="POST">
            <input type="password" name="clave" placeholder="Ingresa la contraseña" required style="padding: 10px; font-size: 16px;">
            <button type="submit" style="padding: 10px 20px; font-size: 16px; cursor: pointer;">Entrar</button>
        </form>
    </body>
    </html>
    """
    return html

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin_autenticado'):
        return redirect(url_for('admin_login'))

    conn = sqlite3.connect('inventario_papichis.db')
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre_producto, categoria, cantidad_actual, precio_venta, precio_compra, ultima_actualizacion FROM productos ORDER BY cantidad_actual ASC")
    datos_brutos = cursor.fetchall()
    conn.close()
# --- LIMPIEZA DE DATOS (ANTI ERROR 500) ---
    datos = []
    for p in datos_brutos:
        id_prod, nombre, cat, cant, p_venta, p_compra, ult_act = p
        
        cant = int(cant) if cant else 0
        
        # Forzar a que el precio de venta sea decimal
        try:
            p_venta = float(p_venta) if p_venta else 0.0
        except ValueError:
            p_venta = 0.0
            
        # NUEVO: Limpiar el precio de compra para evitar errores
        try:
            p_compra = float(p_compra) if p_compra else 0.0
        except ValueError:
            p_compra = 0.0
            
        datos.append((id_prod, nombre, cat, cant, p_venta, p_compra, ult_act))

    # --- ZONA DE CÁLCULOS ANALÍTICOS ---
    total_items = len(datos)
    total_unidades = sum(p[3] for p in datos)
    items_criticos = sum(1 for p in datos if p[3] <= 5)
    valor_total_mercancia = sum(p[3] * p[4] for p in datos)
    
    # NUEVO: Cálculo automático de la Inversión (Cantidad * Precio Compra)
    inversion_total = sum(p[3] * p[5] for p in datos)
   
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Dashboard Analítico Papichis</title>
        <style>
            body { font-family: Arial, sans-serif; padding: 20px; background: #f8f9fa; color: #333; margin: 0; }
            .encabezado { display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid #1f497d; padding-bottom: 15px; margin-bottom: 20px; }
            h2 { margin: 0; color: #1f497d; }
            .btn-excel { background-color: #28a745; color: white; text-decoration: none; padding: 12px 24px; border-radius: 6px; font-weight: bold; font-size: 14px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
            .btn-excel:hover { background-color: #218838; }
            
            .contenedor-kpis { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 20px; margin-bottom: 25px; }
            .tarjeta-kpi { background: white; padding: 20px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.05); border-left: 5px solid #1f497d; }
            .tarjeta-kpi.alerta { border-left-color: #dc3545; }
            .tarjeta-kpi.dinero { border-left-color: #28a745; }
            .kpi-titulo { font-size: 12px; text-transform: uppercase; color: #777; font-weight: bold; margin-bottom: 5px; }
            .kpi-valor { font-size: 24px; font-weight: bold; color: #222; }
            
            table { width: 100%; border-collapse: collapse; background: white; box-shadow: 0 4px 6px rgba(0,0,0,0.05); border-radius: 8px; overflow: hidden; }
            th, td { padding: 14px; text-align: left; border-bottom: 1px solid #eee; vertical-align: middle; }
            th { background-color: #1f497d; color: white; font-weight: bold; }
            tr:hover { background-color: #f1f3f5; }
            .agotado { background-color: #ffe6e6; color: #cc0000; font-weight: bold; }
            
            .form-precio { display: flex; gap: 4px; align-items: center; margin: 0; }
            .input-precio { width: 75px; padding: 6px; font-size: 14px; border: 1px solid #ccc; border-radius: 4px; text-align: center; }
            .btn-guardar-precio { background: #28a745; color: white; border: none; padding: 6px 10px; border-radius: 4px; font-weight: bold; cursor: pointer; }
            .btn-guardar-precio:hover { background: #218838; }
        </style>
    </head>
    <body>
        <div class="encabezado">
            <h2>📊 Panel de Inteligencia de Negocio - Papichis</h2>
            <a href="/descargar_excel" class="btn-excel">📥 Descargar Reporte Financiero (Excel)</a>
        </div>
        
        <div class="contenedor-kpis">
            <div class="tarjeta-kpi">
                <div class="kpi-titulo">Modelos Registrados</div>
                <div class="kpi-valor">{{ total_items }}</div>
            </div>
            <div class="tarjeta-kpi">
                <div class="kpi-titulo">Total Unidades Físicas</div>
                <div class="kpi-valor">{{ total_unidades }}</div>
            </div>
            <div class="tarjeta-kpi dinero">
                <div class="kpi-titulo">Valor Estimado de Venta</div>
                <div class="kpi-valor">${{ "{:,.2f}".format(valor_total_mercancia) }}</div>
            </div>
            
            <!-- NUEVA TARJETA: INVERSIÓN TOTAL -->
            <div class="tarjeta-kpi" style="border-left-color: #ffc107;">
                <div class="kpi-titulo">Precio Stock Invertido</div>
                <div class="kpi-valor">${{ "{:,.2f}".format(inversion_total) }}</div>
            </div>
            
            <div class="tarjeta-kpi alerta">
                <div class="kpi-titulo">Alertas de Reposición</div>
                <div class="kpi-valor" style="color: {% if items_criticos > 0 %}#dc3545{% else %}#222{% endif %};">{{ items_criticos }}</div>
            </div>
        </div>

        <p style="color: #555; margin-bottom: 15px;">Auditoría de almacén en tiempo real. Los valores se recalculan dinámicamente con cada cambio.</p>
        
        <table>
            <thead>
                <tr>
                    <th>Producto</th>
                    <th>Categoría</th>
                    <th>Stock Actual</th>
                    <th>Precio de Venta</th>
                    <th>Valor en Stock</th>
                    <th>Última Revisión</th>
                </tr>
            </thead>
            <tbody>
                {% for producto in datos %}
                <tr class="{% if producto[3] <= 5 %}agotado{% endif %}">
                    <td>{{ producto[1] }}</td>
                    <td>{{ producto[2] }}</td>
                    <td>{{ producto[3] }} uds</td>
                    <td>
                        <form action="/admin/actualizar_precio" method="POST" class="form-precio">
                            <input type="hidden" name="id_producto" value="{{ producto[0] }}">
                            <span style="color: #28a745; font-weight: bold;">$</span>
                            <input type="number" name="nuevo_precio" value="{{ producto[4] }}" step="0.01" min="0" class="input-precio">
                            <button type="submit" class="btn-guardar-precio" title="Guardar Precio">✓</button>
                        </form>
                    </td>
                    <td>${{ "{:,.2f}".format(producto[3] * producto[4]) }}</td>
                    <td>{{ producto[6] or 'No revisado aún' }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        <br>
        <div style="display: flex; justify-content: space-between;">
            <span style="font-size: 12px; color: #777;">* Nota: El margen neto se habilitará al ingresar costos de proveedores.</span>
            <a href="/admin/logout" style="color: #dc3545; text-decoration: none; font-weight: bold; font-size: 14px;">Cerrar Sesión</a>
        </div>
    </body>
    </html>
    """
    return render_template_string(html, datos=datos, total_items=total_items, total_unidades=total_unidades, valor_total_mercancia=valor_total_mercancia, inversion_total=inversion_total, items_criticos=items_criticos)

# --- NUEVA RUTA: ACTUALIZACIÓN DE PRECIOS DESDE EL DASHBOARD ---

@app.route('/admin/actualizar_precio', methods=['POST'])
def admin_actualizar_precio():
    if not session.get('admin_autenticado'):
        return redirect(url_for('admin_login'))
        
    producto_id = request.form['id_producto']
    nuevo_precio_str = request.form['nuevo_precio']
    
    # Limpieza: Si el input viene vacío, lo convertimos a 0.0 para no romper la BD
    try:
        nuevo_precio = float(nuevo_precio_str)
    except ValueError:
        nuevo_precio = 0.0

    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    conn = sqlite3.connect('inventario_papichis.db')
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE productos 
        SET precio_venta = ?, ultima_actualizacion = ? 
        WHERE id = ?
    ''', (nuevo_precio, fecha_actual, producto_id))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))


# --- RUTA DE GENERACIÓN EN TIEMPO REAL DEL EXCEL FINANCIERO ---

@app.route('/descargar_excel')
def descargar_excel():
    if not session.get('admin_autenticado'):
        return redirect(url_for('admin_login'))

    conn = sqlite3.connect('inventario_papichis.db')
    cursor = conn.cursor()
    cursor.execute("SELECT id, categoria, subcategoria, nombre_producto, cantidad_actual, precio_venta, precio_compra, ultima_actualizacion FROM productos")
    productos_bd_brutos = cursor.fetchall()
    conn.close()

    # Limpieza también para el Excel
    productos_bd = []
    for f in productos_bd_brutos:
        id_prod, cat, subcat, name, qty, p_venta, p_compra, last_rev = f
        qty = int(qty) if qty else 0
        try:
            p_venta = float(p_venta) if p_venta else 0.0
        except ValueError:
            p_venta = 0.0
        productos_bd.append((id_prod, cat, subcat, name, qty, p_venta, p_compra, last_rev))

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen Financiero"
    ws_data = wb.create_sheet(title="Inventario Valorizado")
    
    ws_summary.views.sheetView[0].showGridLines = True
    ws_data.views.sheetView[0].showGridLines = True

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_sub_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    fill_alert = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    
    font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_alert = Font(name="Calibri", size=11, bold=True, color="C00000")

    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    border_total = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))

    ws_data.append([])
    ws_data.append(["REPORTE FINANCIERO DE INVENTARIO VALORIZADO"])
    ws_data.cell(row=2, column=1).font = font_title
    ws_data.append([])
    
    headers = ["ID", "Categoría", "Subcategoría", "Producto", "Stock", "P. Venta", "Valor Inventario", "Estado Alerta", "Última Revisión"]
    ws_data.append(headers)
    
    for col_num, h in enumerate(headers, 1):
        c = ws_data.cell(row=4, column=col_num)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal='center' if col_num in [1,5,6,7,8,9] else 'left', vertical='center')

    start_row = 5
    for idx, fila in enumerate(productos_bd, 1):
        r_num = start_row + idx - 1
        _, cat, subcat, name, qty, p_venta, p_compra, last_rev = fila
        
        estado_alerta = "CRÍTICO" if qty <= 5 else "OK"
        valor_stock = qty * p_venta
        revision_texto = last_rev if last_rev else datetime.now().strftime("%Y-%m-%d %H:%M")
        
        ws_data.append([idx, cat, subcat, name, qty, p_venta, valor_stock, estado_alerta, revision_texto])
        
        for col_idx in range(1, 10):
            cell = ws_data.cell(row=r_num, column=col_idx)
            cell.font = font_regular
            cell.border = border_thin
            if r_num % 2 == 0:
                cell.fill = fill_zebra
            
            cell.alignment = Alignment(horizontal='center' if col_idx in [1,5,6,7,8,9] else 'left', vertical='center')
            
            if col_idx == 5:
                cell.number_format = '#,##0'
            elif col_idx in [6, 7]:
                cell.number_format = '$#,##0.00'

    tot_row = start_row + len(productos_bd)
    ws_data.cell(row=tot_row, column=4, value="Totales Consolidados").font = font_bold
    ws_data.cell(row=tot_row, column=4).alignment = Alignment(horizontal='right')
    
    suma_total_stock = sum(f[4] for f in productos_bd)
    suma_total_dinero = sum(f[4] * f[5] for f in productos_bd)
    items_criticos = sum(1 for f in productos_bd if f[4] <= 5)
    
    t_stock = ws_data.cell(row=tot_row, column=5, value=suma_total_stock)
    t_stock.font = font_bold
    t_stock.border = border_total
    t_stock.number_format = '#,##0'
    
    t_dinero = ws_data.cell(row=tot_row, column=7, value=suma_total_dinero)
    t_dinero.font = font_bold
    t_dinero.border = border_total
    t_dinero.number_format = '$#,##0.00'

    rule = CellIsRule(operator='equal', formula=['"CRÍTICO"'], fill=fill_alert, font=font_alert)
    ws_data.conditional_formatting.add(f"H5:H{tot_row-1}", rule)

    ws_summary.append([])
    ws_summary.append(["SISTEMA DE CONTROL DE INVENTARIO - PAPICHIS"])
    ws_summary.cell(row=2, column=2).font = font_title
    ws_summary.append([])
    ws_summary.append(["", "ESTADOS GENERALES Y PROYECCIÓN DE FLUJO"])
    ws_summary.cell(row=4, column=2).font = Font(name="Calibri", size=11, bold=True)
    ws_summary.cell(row=4, column=2).fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    ws_summary.append([])
    
    ws_summary.cell(row=6, column=2, value="Indicador Financiero / Logístico").font = font_bold
    ws_summary.cell(row=6, column=2).fill = fill_sub_header
    ws_summary.cell(row=6, column=2).border = border_thin
    ws_summary.cell(row=6, column=3, value="Valor").font = font_bold
    ws_summary.cell(row=6, column=3).fill = fill_sub_header
    ws_summary.cell(row=6, column=3).border = border_thin
    ws_summary.cell(row=6, column=3).alignment = Alignment(horizontal='center')

    metrics = [
        ("Modelos Únicos Registrados", len(productos_bd)),
        ("Volumen Total de Unidades en Almacén", suma_total_stock),
        ("Capital Estimado de Retorno (Venta Total)", suma_total_dinero),
        ("Alertas de Abastecimiento Crítico (<=5)", items_criticos)
    ]

    for r_idx, (m_name, m_val) in enumerate(metrics, 7):
        c_name = ws_summary.cell(row=r_idx, column=2, value=m_name)
        c_name.font = font_regular
        c_name.border = border_thin
        
        c_val = ws_summary.cell(row=r_idx, column=3, value=m_val)
        c_val.font = font_bold
        c_val.border = border_thin
        
        if r_idx == 9:
            c_val.number_format = '$#,##0.00'
            c_val.alignment = Alignment(horizontal='right')
        else:
            c_val.alignment = Alignment(horizontal='center')
            
        if r_idx == 10 and items_criticos > 0:
            c_name.fill = fill_alert
            c_name.font = font_alert
            c_val.fill = fill_alert
            c_val.font = font_alert

    for ws in [ws_summary, ws_data]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    ws_summary.column_dimensions['A'].width = 3
    ws_summary.column_dimensions['B'].width = 40
    ws_summary.column_dimensions['C'].width = 20

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Reporte_Financiero_Papichis_{datetime.now().strftime("%d_%m_%Y")}.xlsx'
    )

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_autenticado', None)
    return redirect(url_for('mostrar_inventario'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)